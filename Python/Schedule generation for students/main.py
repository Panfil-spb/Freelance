import openpyxl
from openpyxl.styles import Alignment
# класс предмета в учебном плане
class Subject:
    __slots__ = 'Name', 'Type', 'Time', 'NumberPairs', 'CountWeek1', 'CountWeek2'
# класс предмета в расписании
class TimeTableSub:
    __slots__ = 'Name', 'Type', 'Flag'
    def __init__(self):
        self.Name = ''
        self.Type = ''
        self.Flag = False
# функция чтения из файла
def ReadFile(Number):
    flag = True
    Subjects = []
    all_count = 0
    # открываем файл
    file = open(f'3_sem_{Number}.txt', encoding='utf-8')
    CountWeek = 0
    # читаем строчки
    for line in file:
        # формируем массивы строк из файла разделенные ;
        new_line = line.split(';')

        # если длина равна 1, то это количество недель в семестре
        if len(new_line) == 1:
            CountWeek = int(new_line[0])
        else:
            # проходимся по выделенным часам
            for i in range(1, 4):
                # если часы выделены то
                if int(new_line[i]) != 0:
                    # формируем объект предмета
                    Sub = Subject()
                    # задаем имя
                    Sub.Name = new_line[0]
                    # задаем время
                    Sub.Time = int(new_line[i])
                    # задаем количество пар в две недели
                    Sub.NumberPairs = Sub.Time / 2
                    Count = int(Sub.NumberPairs / CountWeek / 0.5)
                    all_count += Count
                    # делаем распределение количества пар по неделям
                    if Count == 3:
                        # если флаг true
                        if flag:
                            Sub.CountWeek1 = 2
                            Sub.CountWeek2 = 1
                        else:
                            Sub.CountWeek1 = 1
                            Sub.CountWeek2 = 2
                        # меняем флаг
                        flag = not flag
                    # если количество четное
                    elif Count == 2:
                        Sub.CountWeek1 = 1
                        Sub.CountWeek2 = 1
                    elif Count == 1:
                        if flag:
                            Sub.CountWeek1 = 1
                            Sub.CountWeek2 = 0
                        else:
                            Sub.CountWeek1 = 0
                            Sub.CountWeek2 = 1
                        flag = not flag
                    elif Count == 4:
                        Sub.CountWeek1 = 2
                        Sub.CountWeek2 = 2
                    # задаем тип пары
                    if i == 1:
                        Sub.Type = 'Лекция'
                    elif i == 2:
                        Sub.Type = 'Практика'
                    else:
                        Sub.Type = 'Лабораторная'
                    Subjects.append(Sub)


    # закрывем файл
    file.close()
    # фозвращаем количество недель и список предметов
    return CountWeek, Subjects

# функция вычисление дня с минимальным количеством пар
def get_min_day(week):
    # формируем массив
    counts = [0 for _ in range(6)]
    # проходимся по расписанию
    for i in range(6):
        for j in range(4):
            # если есть пара то увеличиваем счетчик
            if week[i][j].Flag:
                counts[i] += 1
    # возвращаем индекс дня с минимальным количеством пар
    return counts.index(min(counts))
# функция формирования таблицы для одной группы
def init_book1(TimeTable, Number):
    # формируем документ
    book = openpyxl.Workbook()
    sheet = book.active
    # устанавливаем ширину ячеек
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 13
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    sheet.merge_cells('A1:B2')
    # формируем ячейки
    sheet.merge_cells('C1:D1')
    sheet.cell(1, 3, "Группа 1")
    sheet.cell(2, 3, "Нечетная")
    sheet.cell(2, 4, "Четная")

    # объединяем нужные ячейки
    for i in range(3, 24, 4):
        sheet.merge_cells(f'A{i}:A{i + 3}')
    # заполняем время
    for i in range(3, 27):
        if i % 4 == 3:
            sheet.cell(i, 2, "9:50-11:20")
        if i % 4 == 0:
            sheet.cell(i, 2, "11:40-13:10")
        if i % 4 == 1:
            sheet.cell(i, 2, "13:40-15:10")
        if i % 4 == 2:
            sheet.cell(i, 2, "15:30-17:00")
    # запоняем нужные ячейки
    sheet.cell(3, 1, "Понедельник")
    sheet.cell(7, 1, "Вторник")
    sheet.cell(11, 1, "Среда")
    sheet.cell(15, 1, "Четверг")
    sheet.cell(19, 1, "Пятница")
    sheet.cell(23, 1, "Суббота")

    # заполняем  ячейки предметами
    for i in range(2):
        # номер столбца
        count = 3
        # выбираеем неделю
        week = TimeTable[i]
        # проходимся по расписанию
        for day in week:
            for sub in day:
                # записываем предмет
                sheet.cell(count, i + 3).alignment = Alignment(wrap_text=True,)
                sheet.cell(count, i + 3).value = sub.Name + "\n" + sub.Type
                count += 1
    # сохраняем файл
    book.save(f'Расписание{Number}.xlsx')

# функция формирования таблицы для двух групп
def init_book(TimeTable, Number):
    # формируем документ
    book = openpyxl.Workbook()
    sheet = book.active
    # устанавливаем ширину ячеек
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 13
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    sheet.merge_cells('A1:B2')
    # формируем ячейки
    sheet.merge_cells('C1:D1')
    sheet.cell(1, 3, "Группа 1")
    sheet.cell(2, 3, "Нечетная")
    sheet.cell(2, 4, "Четная")
    # формируем ячейки
    sheet.merge_cells('E1:F1')
    sheet.cell(1, 5, "Группа 2")
    sheet.cell(2, 5, "Нечетная")
    sheet.cell(2, 6, "Четная")

    # объединяем нужные ячейки
    for i in range(3, 24, 4):
        sheet.merge_cells(f'A{i}:A{i + 3}')
    # заполняем время
    for i in range(3, 27):
        if i % 4 == 3:
            sheet.cell(i, 2, "9:50-11:20")
        if i % 4 == 0:
            sheet.cell(i, 2, "11:40-13:10")
        if i % 4 == 1:
            sheet.cell(i, 2, "13:40-15:10")
        if i % 4 == 2:
            sheet.cell(i, 2, "15:30-17:00")
    # запоняем нужные ячейки
    sheet.cell(3, 1, "Понедельник")
    sheet.cell(7, 1, "Вторник")
    sheet.cell(11, 1, "Среда")
    sheet.cell(15, 1, "Четверг")
    sheet.cell(19, 1, "Пятница")
    sheet.cell(23, 1, "Суббота")

    # заполняем  ячейки предметами
    for i in range(4):
        # номер столбца
        count = 3
        # выбираеем неделю
        week = TimeTable[i]
        # проходимся по расписанию
        for day in week:
            for sub in day:
                # записываем предмет
                sheet.cell(count, i + 3).alignment = Alignment(wrap_text=True,)
                sheet.cell(count, i + 3).value = sub.Name + "\n" + sub.Type
                count += 1
    # сохраняем файл
    book.save(f'Расписание{Number}.xlsx')

# функция формирования расписания для одной группы
def init_timetable1(Subjects):
    # Формирование недель
    week11 = [[TimeTableSub() for i in range(4)] for i in range(6)]
    week12 = [[TimeTableSub() for i in range(4)] for i in range(6)]

    count1 = 0
    # проходимся по предметам
    for i in Subjects:
        count1 += 1
        flag = True
        # если это лекция
        if i.Type == 'Лекция':
            # если предмет есть в двух неделях
            if i.CountWeek1 and i.CountWeek2:
                flag = True
                # начинаем проходится начиная со дня где меньше всего пар
                for k in range(get_min_day(week11), len(week11)):
                    for j in range(len(week11[k])):
                        # если время свободно у двух групп в этот день
                        if not week11[k][j].Flag and not week12[k][j].Flag and flag:
                            Sub = TimeTableSub()
                            Sub.Name = i.Name
                            Sub.Type = i.Type
                            Sub.Flag = True
                            week11[k][j] = Sub
                            week12[k][j] = Sub
                            flag = False
                        if not flag:
                            break
                    if not flag:
                        break
                # если в первую неделю две пары
                if i.CountWeek1 == 2:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week11), len(week11)):
                        for j in range(len(week11[k])):
                            # если время свободно у двух групп в этот день
                            if not week11[k][j].Flag  and not week12[k][j].Flag and flag:
                                Sub = TimeTableSub()
                                Sub.Name = i.Name
                                Sub.Type = i.Type
                                Sub.Flag = True
                                week11[k][j] = Sub
                                week12[k][j] = Sub
                                flag = False
                            if not flag:
                                break
                        if not flag:
                            break
                # если во вторую неделю две пары
                elif i.CountWeek2 == 2:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week12), len(week12)):
                        for j in range(len(week12[k])):
                            # если время свободно у двух групп в этот день
                            if not week11[k][j].Flag and not week12[k][j].Flag and flag:
                                Sub = TimeTableSub()
                                Sub.Name = i.Name
                                Sub.Type = i.Type
                                Sub.Flag = True
                                week11[k][j] = Sub
                                week12[k][j] = Sub
                                flag = False
                            if not flag:
                                break
                        if not flag:
                            break
            else:
                # если пара в первую неделю
                if i.CountWeek1 == 1:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week11), len(week11)):
                        for j in range(len(week11[k])):
                            # если время свободно у двух групп в этот день
                            if not week11[k][j].Flag and flag:
                                Sub = TimeTableSub()
                                Sub.Name = i.Name
                                Sub.Type = i.Type
                                Sub.Flag = True
                                week11[k][j] = Sub
                                flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
                # если пара во вторую неделю
                elif i.CountWeek2 == 1:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week12), len(week12)):
                        for j in range(len(week12[k])):
                            # если время свободно у двух групп в этот день
                            if not week12[k][j].Flag and flag:
                                Sub = TimeTableSub()
                                Sub.Name = i.Name
                                Sub.Type = i.Type
                                Sub.Flag = True
                                week12[k][j] = Sub
                                flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
        else:
            if i.CountWeek1 and i.CountWeek2:
                # заполняем предмет для первой группы
                flag = True
                # начинаем проходится начиная со дня где меньше всего пар
                for k in range(get_min_day(week11), len(week11)):
                    for j in range(len(week11[k])):
                        # если время свободно у двух групп в этот день
                        if not week11[k][j].Flag and flag and not week12[k][j].Flag:
                            Sub = TimeTableSub()
                            Sub.Name = i.Name
                            Sub.Type = i.Type
                            Sub.Flag = True
                            week11[k][j] = Sub
                            week12[k][j] = Sub
                            flag = False
                        elif not flag:
                            break
                    if not flag:
                        break
                # если две пары в первую неделю
                if i.CountWeek1 == 2:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week11), len(week11)):
                        for j in range(len(week11[k])):
                            # если время свободно у двух групп в этот день
                            if not week11[k][j].Flag and flag:
                                Sub = TimeTableSub()
                                Sub.Name = i.Name
                                Sub.Type = i.Type
                                Sub.Flag = True
                                week11[k][j] = Sub
                                flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
                # если две пары во вторую неделю
                elif i.CountWeek2 == 2:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week12), len(week12)):
                        for j in range(len(week12[k])):
                            # если время свободно у двух групп в этот день
                            if not week12[k][j].Flag and flag:
                                Sub = TimeTableSub()
                                Sub.Name = i.Name
                                Sub.Type = i.Type
                                Sub.Flag = True
                                week12[k][j] = Sub
                                flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
            else:
                # если пара в первую неделю
                if i.CountWeek1 == 1:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week11), len(week11)):
                        for j in range(len(week11[k])):
                            # если время свободно у двух групп в этот день
                            if not week11[k][j].Flag and flag:
                                Sub = TimeTableSub()
                                Sub.Name = i.Name
                                Sub.Type = i.Type
                                Sub.Flag = True
                                week11[k][j] = Sub
                                flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
                # если пара во вторую неделю
                elif i.CountWeek2 == 1:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week12), len(week12)):
                        for j in range(len(week12[k])):
                            # если время свободно у двух групп в этот день
                            if not week12[k][j].Flag and flag:
                                Sub = TimeTableSub()
                                Sub.Name = i.Name
                                Sub.Type = i.Type
                                Sub.Flag = True
                                week12[k][j] = Sub
                                flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
    #  возвращаем сформированные недели
    return [week11, week12]

def init_timetable(Subjects):
    # Формирование недель
    week11 = [[TimeTableSub() for i in range(4)] for i in range(6)]
    week12 = [[TimeTableSub() for i in range(4)] for i in range(6)]
    week21 = [[TimeTableSub() for i in range(4)] for i in range(6)]
    week22 = [[TimeTableSub() for i in range(4)] for i in range(6)]

    count1 = 0
    for i in Subjects:
        count1 += 1
        flag = True
        # если это лекция
        if i.Type == 'Лекция':
            if i.CountWeek1 and i.CountWeek2:
                flag = True
                # начинаем проходится начиная со дня где меньше всего пар
                for k in range(get_min_day(week11), len(week11)):
                    for j in range(len(week11[k])):
                        # если время свободно у двух групп в этот день
                        if not week11[k][j].Flag and not week21[k][j].Flag and not week22[k][j].Flag and not week12[k][j].Flag and flag:
                            Sub = TimeTableSub()
                            Sub.Name = i.Name
                            Sub.Type = i.Type
                            Sub.Flag = True
                            week11[k][j] = Sub
                            week12[k][j] = Sub
                            week21[k][j] = Sub
                            week22[k][j] = Sub
                            flag = False
                        if not flag:
                            break
                    if not flag:
                        break

                if i.CountWeek1 == 2:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week11), len(week11)):
                        for j in range(len(week11[k])):
                            # если время свободно у двух групп в этот день
                            if not week11[k][j].Flag and not week21[k][j].Flag and not week22[k][j].Flag and not \
                            week12[k][j].Flag and flag:
                                Sub = TimeTableSub()
                                Sub.Name = i.Name
                                Sub.Type = i.Type
                                Sub.Flag = True
                                week11[k][j] = Sub
                                week12[k][j] = Sub
                                week21[k][j] = Sub
                                week22[k][j] = Sub
                                flag = False
                            if not flag:
                                break
                        if not flag:
                            break
                elif i.CountWeek2 == 2:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week12), len(week12)):
                        for j in range(len(week12[k])):
                            # если время свободно у двух групп в этот день
                            if not week11[k][j].Flag and not week21[k][j].Flag and not week22[k][j].Flag and not \
                            week12[k][j].Flag and flag:
                                Sub = TimeTableSub()
                                Sub.Name = i.Name
                                Sub.Type = i.Type
                                Sub.Flag = True
                                week11[k][j] = Sub
                                week12[k][j] = Sub
                                week21[k][j] = Sub
                                week22[k][j] = Sub
                                flag = False
                            if not flag:
                                break
                        if not flag:
                            break
            else:
                if i.CountWeek1 == 1:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week11), len(week11)):
                        for j in range(len(week11[k])):
                            # если время свободно у двух групп в этот день
                            if not week11[k][j].Flag and not week21[k][j].Flag and flag:
                                Sub = TimeTableSub()
                                Sub.Name = i.Name
                                Sub.Type = i.Type
                                Sub.Flag = True
                                week11[k][j] = Sub
                                week21[k][j] = Sub

                                flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
                elif i.CountWeek2 == 1:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week12), len(week12)):
                        for j in range(len(week12[k])):
                            # если время свободно у двух групп в этот день
                            if not week12[k][j].Flag and not week22[k][j].Flag and flag:
                                Sub = TimeTableSub()
                                Sub.Name = i.Name
                                Sub.Type = i.Type
                                Sub.Flag = True
                                week12[k][j] = Sub
                                week22[k][j] = Sub

                                flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
        # если практика или лаба
        else:
            # если пары каждую неделю
            if i.CountWeek1 and i.CountWeek2:
                # заполняем предмет для первой группы
                flag = True
                # начинаем проходится начиная со дня где меньше всего пар
                for k in range(get_min_day(week11), len(week11)):
                    for j in range(len(week11[k])):
                        # если время свободно у двух групп в этот день
                        if not week11[k][j].Flag and flag and not week12[k][j].Flag:
                            if (not week21[k][j].Flag and not week22[k][j].Flag) or (week21[k][j].Name != i.Name and week22[k][j].Name != i.Name):
                                Sub = TimeTableSub()
                                Sub.Name = i.Name
                                Sub.Type = i.Type
                                Sub.Flag = True
                                week11[k][j] = Sub
                                week12[k][j] = Sub
                                flag = False
                        elif not flag:
                            break
                    if not flag:
                        break
                # заполняем этот предмет для второй группы
                flag = True
                # начинаем проходится начиная со дня где меньше всего пар
                for k in range(get_min_day(week21), len(week21)):
                    for j in range(len(week21[k])):
                        # если время свободно у двух групп в этот день
                        if not week21[k][j].Flag and flag and not week22[k][j].Flag:
                            if (not week11[k][j].Flag and not week12[k][j].Flag) or (
                                    week11[k][j].Name != i.Name and week12[k][j].Name != i.Name):
                                Sub = TimeTableSub()
                                Sub.Name = i.Name
                                Sub.Type = i.Type
                                Sub.Flag = True
                                week21[k][j] = Sub
                                week22[k][j] = Sub
                                flag = False
                        elif not flag:
                            break
                    if not flag:
                        break
                # если две пары в первую неделю
                if i.CountWeek1 == 2:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week11), len(week11)):
                        for j in range(len(week11[k])):
                            # если время свободно у двух групп в этот день
                            if not week11[k][j].Flag and flag:
                                if (not week21[k][j].Flag) or (week21[k][j].Name != i.Name):
                                    Sub = TimeTableSub()
                                    Sub.Name = i.Name
                                    Sub.Type = i.Type
                                    Sub.Flag = True
                                    week11[k][j] = Sub
                                    flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
                    # заполняем этот предмет для второй группы
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week21), len(week21)):
                        for j in range(len(week21[k])):
                            # если время свободно у двух групп в этот день
                            if not week21[k][j].Flag and flag:
                                if (not week11[k][j].Flag) or (week11[k][j].Name != i.Name):
                                    Sub = TimeTableSub()
                                    Sub.Name = i.Name
                                    Sub.Type = i.Type
                                    Sub.Flag = True
                                    week21[k][j] = Sub
                                    flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
                # если две пары во вторую неделю
                elif i.CountWeek2 == 2:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week12), len(week12)):
                        for j in range(len(week12[k])):
                            # если время свободно у двух групп в этот день
                            if not week12[k][j].Flag and flag:
                                if (not week22[k][j].Flag) or (week22[k][j].Name != i.Name):
                                    Sub = TimeTableSub()
                                    Sub.Name = i.Name
                                    Sub.Type = i.Type
                                    Sub.Flag = True
                                    week12[k][j] = Sub
                                    flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
                    # заполняем этот предмет для второй группы
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week22), len(week22)):
                        for j in range(len(week22[k])):
                            # если время свободно у двух групп в этот день
                            if flag and not week22[k][j].Flag:
                                if (not week12[k][j].Flag) or (week12[k][j].Name != i.Name):
                                    Sub = TimeTableSub()
                                    Sub.Name = i.Name
                                    Sub.Type = i.Type
                                    Sub.Flag = True
                                    week22[k][j] = Sub
                                    flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
            # иначе если пара только в одну из недель
            else:
                # если пара в первую неделю
                if i.CountWeek1 == 1:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week11), len(week11)):
                        for j in range(len(week11[k])):
                            # если время свободно у двух групп в этот день
                            if not week11[k][j].Flag and flag:
                                if (not week21[k][j].Flag) or (week21[k][j].Name != i.Name):
                                    Sub = TimeTableSub()
                                    Sub.Name = i.Name
                                    Sub.Type = i.Type
                                    Sub.Flag = True
                                    week11[k][j] = Sub
                                    flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
                    # заполняем этот предмет для второй группы
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week21), len(week21)):
                        for j in range(len(week21[k])):
                            # если время свободно у двух групп в этот день
                            if not week21[k][j].Flag and flag:
                                if (not week11[k][j].Flag) or (week11[k][j].Name != i.Name):
                                    Sub = TimeTableSub()
                                    Sub.Name = i.Name
                                    Sub.Type = i.Type
                                    Sub.Flag = True
                                    week21[k][j] = Sub
                                    flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
                # если пара во вторую неделю
                elif i.CountWeek2 == 1:
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week12), len(week12)):
                        for j in range(len(week12[k])):
                            # если время свободно у двух групп в этот день
                            if not week12[k][j].Flag and flag:
                                if (not week22[k][j].Flag) or (week22[k][j].Name != i.Name):
                                    Sub = TimeTableSub()
                                    Sub.Name = i.Name
                                    Sub.Type = i.Type
                                    Sub.Flag = True
                                    week12[k][j] = Sub
                                    flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
                    # заполняем этот предмет для второй группы
                    flag = True
                    # начинаем проходится начиная со дня где меньше всего пар
                    for k in range(get_min_day(week22), len(week22)):
                        for j in range(len(week22[k])):
                            # если время свободно у двух групп в этот день
                            if flag and not week22[k][j].Flag:
                                if (not week12[k][j].Flag) or (week12[k][j].Name != i.Name):
                                    Sub = TimeTableSub()
                                    Sub.Name = i.Name
                                    Sub.Type = i.Type
                                    Sub.Flag = True
                                    week22[k][j] = Sub
                                    flag = False
                            elif not flag:
                                break
                        if not flag:
                            break
    # возвращаем сформированные недели
    return [week11, week12, week21, week22]




def main():
    for i in range(1, 3):
        # получаем количество недель и список предметов
        CountWeek, Subjects = ReadFile(i)
        # если это первый поток с двумя группами
        if i == 1:
            # формируем расписание
            timetable = init_timetable(Subjects)
            # формируем файл
            init_book(timetable, i)
        # иначе если это поток с одной группой
        else:
            # формируем расписание
            timetable = init_timetable1(Subjects)
            # формируем файл
            init_book1(timetable, i)


if __name__ == '__main__':
    main()